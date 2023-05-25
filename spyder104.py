import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import re

def data_clean(text):
    # 清洗excel中的非法字符，都是不常见的不可显示字符，例如退格，响铃等
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
    text = ILLEGAL_CHARACTERS_RE.sub(r"", text)
    return text

#抓取並解析網頁資料
url ="https://www.104.com.tw/jobs/search/?keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&order=1&jobsource=2018indexpoc&ro=0"
res = requests.get(url)
soup = BeautifulSoup(res.text)
page = 1

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "職位名稱"
ws["B1"] = "公司名稱"
ws["C1"] = "工作地區"
ws["D1"] = "職缺薪資"
ws["E1"] = "職缺連結"

while soup.find_all('article',class_="b-block--top-bord job-list-item b-clearfix js-job-item") !=[]:
    print("===================================================")
    print(f"現在正在讀取{page}頁...")
    print("===================================================")
    #針對最小資造<job>組成資料集中的個別標籤，<article>，做同一件事情
    for job in soup.find_all('article',class_="b-block--top-bord job-list-item b-clearfix js-job-item"):
        #印出5個目標欄位
        jobName = job.a.text#職缺名稱
        company = job.select('li')[1].a.text.strip()#公司名稱
        area = job.find('ul',class_="b-list-inline b-clearfix job-list-intro b-content").li.text#工作地區
        #如果span標籤不存在<搜索span標籤回傳空list>，則印出a.text，否則印出span.text待遇面議
        if job.find('div',class_="job-list-tag b-content").select('span')==[]:#薪資
            salary = job.find('div',class_="job-list-tag b-content").a.text
        else:
            salary = job.find('div',class_="job-list-tag b-content").span.text
        web = "https:"+job.a['href']#連結
        
        ws.append([data_clean(jobName),company,area,salary,web])
        
        
    page+=1
    url = f"https://www.104.com.tw/jobs/search/?ro=0&kwop=7&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&order=12&asc=0&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1&page={str(page)}"
    res = requests.get(url)
    soup = BeautifulSoup(res.text)
    wb.save("104大數據05221.xlsx")
    time.sleep(1)