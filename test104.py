import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "職位名稱"
ws["B1"] = "公司名稱"
ws["C1"] = "工作地區"
ws["D1"] = "職缺薪資"
ws["E1"] = "職缺連結"


ws.append([1,2,3,4,5])

wb.save("testSpyder.xlsx")